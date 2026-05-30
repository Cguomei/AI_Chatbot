# ==================== 启动崩溃日志（最优先，在一切import之前） ====================
import os as _os, sys as _sys, traceback as _tb, datetime as _dt

def _crash_log(msg):
    try:
        if getattr(_sys, "frozen", False):
            d = _os.path.dirname(_sys.executable)
        else:
            d = _os.path.dirname(_os.path.abspath(__file__))
        with open(_os.path.join(d, "crash.log"), "a", encoding="utf-8") as f:
            f.write(f"{_dt.datetime.now().isoformat()} {msg}\n")
            f.flush()
    except:
        pass

try:
    _crash_log("=== STARTUP BEGIN ===")
    _crash_log(f"frozen={getattr(_sys, 'frozen', False)}, argv={_sys.argv}, cwd={_os.getcwd()}")
except:
    pass

# 全局未捕获异常 => 写入 crash.log
_orig_hook = _sys.excepthook
def _crash_hook(typ, val, tb):
    _crash_log(f"UNHANDLED: {typ.__name__}: {val}")
    _crash_log("".join(_tb.format_tb(tb)))
    if _orig_hook:
        _orig_hook(typ, val, tb)
_sys.excepthook = _crash_hook

import hashlib, csv, io, math, os, glob, shutil, logging, sys, time, traceback
_crash_log("step1: stdlib imports OK")
from contextlib import asynccontextmanager
from datetime import datetime
from typing import Optional
from fastapi import FastAPI, Request, Form, HTTPException, Depends, Query
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
_crash_log("step2: fastapi imports OK")
from starlette.middleware.sessions import SessionMiddleware
_crash_log("step3: starlette OK")
from jinja2 import Environment, FileSystemLoader
_crash_log("step4: jinja2 OK")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
_crash_log("step5: openpyxl OK")
from database import get_db, init_db
_crash_log("step6: database OK")

# ==================== 路径兼容（打包成 exe 后资源在 sys._MEIPASS，数据在 exe 同目录）========================
if getattr(sys, 'frozen', False):
    _RESOURCE_DIR = sys._MEIPASS
    _DATA_DIR = os.path.dirname(sys.executable)
else:
    _RESOURCE_DIR = os.path.dirname(__file__)
    _DATA_DIR = _RESOURCE_DIR
_crash_log(f"step7: path OK, frozen={getattr(sys, 'frozen', False)}, resource={_RESOURCE_DIR}")

# ==================== 日志配置 ====================
LOG_DIR = os.path.join(_DATA_DIR, "logs")
os.makedirs(LOG_DIR, exist_ok=True)
_crash_log(f"step8: log dir created: {LOG_DIR}")

logger = logging.getLogger("boardgame")
logger.setLevel(logging.DEBUG)

# 控制台 handler
ch = logging.StreamHandler(sys.stderr)
ch.setLevel(logging.INFO)
ch.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S"))
logger.addHandler(ch)

# 文件 handler（所有级别都写入）
fh = logging.FileHandler(os.path.join(LOG_DIR, "server.log"), encoding="utf-8")
fh.setLevel(logging.DEBUG)
fh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
logger.addHandler(fh)

_TEMPLATE_DIR = os.path.join(_RESOURCE_DIR, "templates")

logger.info("=" * 50)
logger.info(f"启动中... Python={sys.version}, 工作目录={os.getcwd()}")
logger.info(f"模板目录={_TEMPLATE_DIR}")

BACKUP_DIR = os.path.join(_DATA_DIR, "db", "backups")
MAX_AUTO_BACKUPS = 10


def ensure_backup_dir():
    os.makedirs(BACKUP_DIR, exist_ok=True)


def auto_backup_db():
    """在记录新对局后自动备份数据库"""
    from database import DB_PATH
    ensure_backup_dir()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(BACKUP_DIR, f"auto_backup_{timestamp}.db")
    shutil.copy2(DB_PATH, backup_path)
    # 清理旧备份，保留最近 MAX_AUTO_BACKUPS 个
    files = sorted(glob.glob(os.path.join(BACKUP_DIR, "auto_backup_*.db")))
    while len(files) > MAX_AUTO_BACKUPS:
        os.remove(files.pop(0))


@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    yield


app = FastAPI(title="桌游排行", lifespan=lifespan)
_crash_log("step9: FastAPI app created")
app.add_middleware(SessionMiddleware, secret_key="boardgame-secret-key-2026")
_crash_log("step10: session middleware OK")

# ==================== 请求日志中间件 ====================
@app.middleware("http")
async def log_requests(request: Request, call_next):
    start = time.time()
    url = str(request.url)
    try:
        response = await call_next(request)
        elapsed = (time.time() - start) * 1000
        logger.info(f"{request.method} {url} -> {response.status_code} ({elapsed:.0f}ms)")
        return response
    except Exception:
        elapsed = (time.time() - start) * 1000
        logger.error(f"{request.method} {url} -> 500 ({elapsed:.0f}ms)\n{traceback.format_exc()}")
        raise

app.mount("/static", StaticFiles(directory=os.path.join(_RESOURCE_DIR, "static")), name="static")
_crash_log("step11: static files mounted")

# 用原始 Jinja2 替代 Starlette 的 Jinja2Templates，绕过 Starlette 0.50 + Jinja2 3.1.6 缓存 key 问题
_jinja_env = Environment(loader=FileSystemLoader(_TEMPLATE_DIR))
_crash_log("step12: jinja2 env OK")


def render_template(name: str, request: Request, context: dict = None) -> HTMLResponse:
    """渲染模板，绕过 Starlette 0.50 的 cache bug"""
    ctx = dict(context) if context else {}
    ctx.setdefault("request", request)
    template = _jinja_env.get_template(name)
    return HTMLResponse(template.render(**ctx))


ELO_K = 32  # ELO K因子


# ==================== Excel 导出样式工具 ====================
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(name="微软雅黑", size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
EVEN_FILL = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")


def style_excel_sheet(ws, headers, col_widths=None):
    """给工作表的表头和内容统一美化"""
    # 表头样式
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # 数据行样式 + 斑马条纹
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)), 2):
        for cell in row:
            cell.font = DATA_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = EVEN_FILL

    # 列宽
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        # 自动宽度
        for col_idx in range(1, len(headers) + 1):
            max_len = len(str(headers[col_idx - 1])) * 2
            for row_idx in range(2, ws.max_row + 1):
                val = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_len = max(max_len, len(val) * 1.5)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    # 冻结首行
    ws.freeze_panes = "A2"


def make_excel_response(wb, filename):
    """将 openpyxl Workbook 转为可下载的 StreamingResponse"""
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()


def get_admin(request: Request):
    return request.session.get("admin")


def require_admin(request: Request):
    admin = get_admin(request)
    if not admin:
        raise HTTPException(status_code=302, headers={"Location": "/admin/login"})
    return admin


def get_active_season(conn):
    cursor = conn.cursor()
    cursor.execute("SELECT id, name FROM seasons WHERE is_active = 1 LIMIT 1")
    return cursor.fetchone()


def get_setting(conn, key, default=None):
    cursor = conn.cursor()
    cursor.execute("SELECT value FROM settings WHERE key = ?", (key,))
    row = cursor.fetchone()
    return row["value"] if row else default


def calc_elo(player_rating, opponent_avg_rating, actual_score):
    """计算新 ELO"""
    expected = 1.0 / (1.0 + math.pow(10, (opponent_avg_rating - player_rating) / 400.0))
    return round(player_rating + ELO_K * (actual_score - expected))


def check_achievements(conn, player_id):
    """检查并发放新成就"""
    cursor = conn.cursor()

    # ===== 全局统计 =====
    cursor.execute("""
        SELECT COUNT(*) FROM match_players mp
        JOIN matches m ON mp.match_id = m.id
        WHERE mp.player_id = ?
    """, (player_id,))
    total_matches = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COUNT(*) FROM match_players mp
        JOIN matches m ON mp.match_id = m.id
        WHERE mp.player_id = ? AND mp.team = m.winner
    """, (player_id,))
    total_wins = cursor.fetchone()[0]

    # 总积分（非ELO）
    cursor.execute("""
        SELECT COALESCE(SUM(mp.score_change), 0)
        FROM match_players mp WHERE mp.player_id = ?
    """, (player_id,))
    total_score = cursor.fetchone()[0]

    # 最高单场得分
    cursor.execute("""
        SELECT COALESCE(MAX(mp.score_change), 0)
        FROM match_players mp WHERE mp.player_id = ?
    """, (player_id,))
    high_score = cursor.fetchone()[0]

    # MVP次数（本局得分最高，不含并列时的重复计数）
    cursor.execute("""
        SELECT COUNT(*) FROM match_players mp
        WHERE mp.player_id = ?
          AND mp.score_change = (SELECT MAX(mp2.score_change) FROM match_players mp2
                                 WHERE mp2.match_id = mp.match_id)
          AND mp.score_change > 0
    """, (player_id,))
    mvp_count = cursor.fetchone()[0]

    # 连胜
    cursor.execute("""
        SELECT mp.team = m.winner as is_win
        FROM match_players mp JOIN matches m ON mp.match_id = m.id
        WHERE mp.player_id = ? ORDER BY m.played_at DESC
    """, (player_id,))
    results = cursor.fetchall()
    consecutive_wins = 0
    consecutive_loses = 0
    for r in results:
        if r[0]:
            consecutive_wins += 1
        else:
            break
    for r in results:
        if not r[0]:
            consecutive_loses += 1
        else:
            break

    # 好人/坏人胜率
    cursor.execute("""
        SELECT COUNT(*), SUM(CASE WHEN mp.team = m.winner THEN 1 ELSE 0 END)
        FROM match_players mp JOIN matches m ON mp.match_id = m.id
        WHERE mp.player_id = ? AND mp.team = 'good'
    """, (player_id,))
    g = cursor.fetchone()
    good_winrate = round(g[1] / g[0] * 100) if g[0] > 0 else 0

    cursor.execute("""
        SELECT COUNT(*), SUM(CASE WHEN mp.team = m.winner THEN 1 ELSE 0 END)
        FROM match_players mp JOIN matches m ON mp.match_id = m.id
        WHERE mp.player_id = ? AND mp.team = 'evil'
    """, (player_id,))
    e = cursor.fetchone()
    evil_winrate = round(e[1] / e[0] * 100) if e[0] > 0 else 0

    cursor.execute("SELECT elo_rating FROM players WHERE id = ?", (player_id,))
    elo = cursor.fetchone()[0]

    # 各游戏统计：{game_id: {"matches": n, "wins": n}}
    cursor.execute("""
        SELECT m.game_id,
               COUNT(*) as matches,
               SUM(CASE WHEN mp.team = m.winner THEN 1 ELSE 0 END) as wins
        FROM match_players mp JOIN matches m ON mp.match_id = m.id
        WHERE mp.player_id = ? GROUP BY m.game_id
    """, (player_id,))
    game_stats = {row[0]: {"matches": row[1], "wins": row[2]} for row in cursor.fetchall()}

    # 各角色统计：{role_id: {"matches": n, "wins": n}}
    cursor.execute("""
        SELECT mp.role_id,
               COUNT(*) as matches,
               SUM(CASE WHEN mp.team = m.winner THEN 1 ELSE 0 END) as wins
        FROM match_players mp JOIN matches m ON mp.match_id = m.id
        WHERE mp.player_id = ? AND mp.role_id IS NOT NULL
        GROUP BY mp.role_id
    """, (player_id,))
    role_stats = {row[0]: {"matches": row[1], "wins": row[2]} for row in cursor.fetchall()}

    global_stats = {
        "total_matches": total_matches,
        "total_wins": total_wins,
        "total_score": total_score,
        "high_score": high_score,
        "mvp_count": mvp_count,
        "consecutive_wins": consecutive_wins,
        "consecutive_loses": consecutive_loses,
        "good_win_rate": good_winrate,
        "evil_win_rate": evil_winrate,
        "elo_rating": elo,
    }

    cursor.execute("SELECT achievement_id FROM player_achievements WHERE player_id = ?", (player_id,))
    earned = {r[0] for r in cursor.fetchall()}

    cursor.execute("SELECT id, name, condition_type, condition_value, game_id, role_id FROM achievements")
    new_achievements = []
    for ach in cursor.fetchall():
        ach_id, name, ctype, cval, gid, rid = ach
        if ach_id in earned:
            continue

        val = None
        if ctype in global_stats:
            # 全局条件
            val = global_stats[ctype]
        elif ctype == "game_matches":
            if gid and gid in game_stats:
                val = game_stats[gid]["matches"]
        elif ctype == "game_wins":
            if gid and gid in game_stats:
                val = game_stats[gid]["wins"]
        elif ctype == "role_matches":
            if rid and rid in role_stats:
                val = role_stats[rid]["matches"]
        elif ctype == "role_wins":
            if rid and rid in role_stats:
                val = role_stats[rid]["wins"]

        if val is not None and val >= cval:
            cursor.execute(
                "INSERT OR IGNORE INTO player_achievements (player_id, achievement_id) VALUES (?, ?)",
                (player_id, ach_id)
            )
            if cursor.rowcount > 0:
                new_achievements.append(name)

    conn.commit()
    return new_achievements


# ==================== 页面路由 ====================

@app.get("/", response_class=HTMLResponse)
async def index(request: Request, game_id: str = "", season_id: str = ""):
    logger.info(f"[index] game_id={game_id!r} season_id={season_id!r}")
    gid = int(game_id) if game_id else None
    sid_raw = int(season_id) if season_id else None
    try:
        conn = get_db()
        cursor = conn.cursor()

        season = get_active_season(conn)
        sid = sid_raw or (season[0] if season else None)

        cursor.execute("SELECT COUNT(*) FROM matches WHERE season_id = ? OR season_id IS NULL", (sid,))
        total_matches = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM players")
        total_players = cursor.fetchone()[0]

        # 最高分
        query = """
            SELECT p.name, COALESCE(SUM(mp.score_change), 0) as total_score
            FROM players p
            LEFT JOIN match_players mp ON p.id = mp.player_id
            LEFT JOIN matches m ON mp.match_id = m.id
            WHERE (? IS NULL OR m.season_id = ? OR m.season_id IS NULL)
            GROUP BY p.id ORDER BY total_score DESC LIMIT 1
        """
        cursor.execute(query, (sid, sid))
        top_row = cursor.fetchone()
        top_score = top_row[1] if top_row else 0

        cursor.execute("SELECT id, name FROM games ORDER BY id")
        games = cursor.fetchall()

        cursor.execute("SELECT id, name FROM seasons ORDER BY started_at DESC")
        seasons = cursor.fetchall()

        # 排行榜
        if gid:
            query = """
                SELECT p.id, p.name, p.avatar, p.elo_rating,
                       COALESCE(SUM(mp.score_change), 0) as total_score,
                       COUNT(mp.id) as match_count,
                       ROUND(CAST(SUM(CASE WHEN (mp.team = m.winner) THEN 1 ELSE 0 END) AS FLOAT) / NULLIF(COUNT(mp.id), 0) * 100, 1) as win_rate
                FROM players p
                LEFT JOIN match_players mp ON p.id = mp.player_id
                LEFT JOIN matches m ON mp.match_id = m.id AND m.game_id = ?
                WHERE mp.id IS NULL OR (m.game_id = ? AND (? IS NULL OR m.season_id = ? OR m.season_id IS NULL))
                GROUP BY p.id ORDER BY total_score DESC
            """
            cursor.execute(query, (gid, gid, sid, sid))
        else:
            query = """
                SELECT p.id, p.name, p.avatar, p.elo_rating,
                       COALESCE(SUM(mp.score_change), 0) as total_score,
                       COUNT(mp.id) as match_count,
                       ROUND(CAST(SUM(CASE WHEN (mp.team = m.winner) THEN 1 ELSE 0 END) AS FLOAT) / NULLIF(COUNT(mp.id), 0) * 100, 1) as win_rate
                FROM players p
                LEFT JOIN match_players mp ON p.id = mp.player_id
                LEFT JOIN matches m ON mp.match_id = m.id
                WHERE mp.id IS NULL OR (? IS NULL OR m.season_id = ? OR m.season_id IS NULL)
                GROUP BY p.id HAVING COUNT(mp.id) > 0
                ORDER BY total_score DESC
            """
            cursor.execute(query, (sid, sid))

        rankings = cursor.fetchall()
        settings = {r["key"]: r["value"] for r in cursor.execute("SELECT key, value FROM settings").fetchall()}
        conn.close()

        return render_template("index.html", request, {
            "total_matches": total_matches, "total_players": total_players,
            "top_score": top_score, "games": games, "rankings": rankings,
            "current_game_id": gid, "current_season_id": sid,
            "seasons": seasons, "season": dict(season) if season else None,
            "admin": get_admin(request), "settings": settings
        })
    except Exception:
        logger.error(f"[index] 崩溃: game_id={game_id!r} season_id={season_id!r}\n{traceback.format_exc()}")
        raise


@app.get("/record", response_class=HTMLResponse)
async def record_page(request: Request):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, name, use_elo, enable_roles, enable_actions FROM games ORDER BY id")
    games = cursor.fetchall()
    cursor.execute("SELECT id, name, avatar, elo_rating FROM players ORDER BY elo_rating DESC")
    players = cursor.fetchall()
    cursor.execute("SELECT id, name, game_id FROM game_presets ORDER BY created_at DESC")
    presets = cursor.fetchall()
    conn.close()
    return render_template("record.html", request, {
        "games": games, "players": players,
        "presets": presets, "admin": get_admin(request)
    })


@app.get("/player/{player_id}", response_class=HTMLResponse)
async def player_detail(request: Request, player_id: int):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM players WHERE id = ?", (player_id,))
    player = cursor.fetchone()
    if not player:
        raise HTTPException(status_code=404)

    cursor.execute("SELECT COALESCE(SUM(score_change), 0) FROM match_players WHERE player_id = ?", (player_id,))
    total_score = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM match_players WHERE player_id = ?", (player_id,))
    total_matches = cursor.fetchone()[0]

    cursor.execute("""
        SELECT ROUND(CAST(SUM(CASE WHEN mp.team = m.winner THEN 1 ELSE 0 END) AS FLOAT) / NULLIF(COUNT(*), 0) * 100, 1)
        FROM match_players mp JOIN matches m ON mp.match_id = m.id WHERE mp.player_id = ?
    """, (player_id,))
    win_rate = cursor.fetchone()[0] or 0

    cursor.execute("SELECT ROUND(CAST(COALESCE(SUM(score_change), 0) AS FLOAT) / NULLIF(COUNT(*), 0), 1) FROM match_players WHERE player_id = ?", (player_id,))
    avg_score = cursor.fetchone()[0] or 0

    cursor.execute("""
        SELECT COUNT(*), ROUND(CAST(SUM(CASE WHEN mp.team = m.winner THEN 1 ELSE 0 END) AS FLOAT) / NULLIF(COUNT(*), 0) * 100, 1),
               ROUND(CAST(COALESCE(SUM(mp.score_change), 0) AS FLOAT) / NULLIF(COUNT(*), 0), 1)
        FROM match_players mp JOIN matches m ON mp.match_id = m.id
        WHERE mp.player_id = ? AND mp.team = 'good'
    """, (player_id,))
    g = cursor.fetchone()

    cursor.execute("""
        SELECT COUNT(*), ROUND(CAST(SUM(CASE WHEN mp.team = m.winner THEN 1 ELSE 0 END) AS FLOAT) / NULLIF(COUNT(*), 0) * 100, 1),
               ROUND(CAST(COALESCE(SUM(mp.score_change), 0) AS FLOAT) / NULLIF(COUNT(*), 0), 1)
        FROM match_players mp JOIN matches m ON mp.match_id = m.id
        WHERE mp.player_id = ? AND mp.team = 'evil'
    """, (player_id,))
    e = cursor.fetchone()

    cursor.execute("""
        SELECT g.name, m.winner, mp.team, mp.score_change, mp.elo_change, m.played_at,
               m.id as match_id, m.game_id,
               COALESCE(r.name, '') as role_name,
               COALESCE(r.score_bonus, 0) as role_bonus,
               COALESCE(g.lose_mode, 'keep_bonus') as lose_mode,
               COALESCE(m.notes, '') as match_notes
        FROM match_players mp JOIN matches m ON mp.match_id = m.id
        JOIN games g ON m.game_id = g.id
        LEFT JOIN roles r ON mp.role_id = r.id
        WHERE mp.player_id = ? ORDER BY m.played_at DESC LIMIT 30
    """, (player_id,))
    history = cursor.fetchall()

    # 查询每局的操作加分
    history_with_actions = []
    for h in history:
        hd = dict(h)
        hd["actions"] = []
        cursor.execute("""
            SELECT ga.name, ga.score_bonus
            FROM match_player_actions mpa
            JOIN game_actions ga ON mpa.action_id = ga.id
            WHERE mpa.match_id = ? AND mpa.player_id = ?
        """, (h["match_id"], player_id))
        hd["actions"] = [dict(a) for a in cursor.fetchall()]
        hd["action_total"] = sum(a["score_bonus"] for a in hd["actions"])

        # 根据 lose_mode 调整失败方显示的加分值，使 base_score = score_change - role_bonus_adj - action_total_adj 始终成立
        lose_mode = hd.get("lose_mode", "keep_bonus")
        role_bonus_raw = hd["role_bonus"]
        action_total_raw = hd["action_total"]
        if hd["team"] != hd["winner"] and lose_mode == "zero":
            hd["role_bonus"] = 0
            hd["action_total"] = 0
            hd["actions"] = []
            hd["base_score"] = 0
        elif hd["team"] != hd["winner"] and lose_mode == "penalize_all":
            hd["role_bonus"] = -role_bonus_raw
            hd["action_total"] = -action_total_raw
            for a in hd["actions"]:
                a["score_bonus"] = -a["score_bonus"]
            hd["base_score"] = hd["score_change"] - hd["role_bonus"] - hd["action_total"]
        else:
            hd["base_score"] = hd["score_change"] - hd["role_bonus"] - hd["action_total"]
        history_with_actions.append(hd)

    # 成就
    cursor.execute("""
        SELECT a.name, a.description, a.icon, pa.earned_at
        FROM player_achievements pa JOIN achievements a ON pa.achievement_id = a.id
        WHERE pa.player_id = ? ORDER BY pa.earned_at DESC
    """, (player_id,))
    achievements = cursor.fetchall()

    # 各角色统计
    cursor.execute("""
        SELECT COALESCE(r.name, '无角色') as role_name, g.name as game_name, mp.team,
               COUNT(*) as times,
               SUM(CASE WHEN mp.team = m.winner THEN 1 ELSE 0 END) as wins,
               ROUND(CAST(COALESCE(AVG(mp.score_change), 0) AS FLOAT), 1) as avg_score
        FROM match_players mp
        JOIN matches m ON mp.match_id = m.id
        JOIN games g ON m.game_id = g.id
        LEFT JOIN roles r ON mp.role_id = r.id
        WHERE mp.player_id = ?
        GROUP BY mp.role_id, g.name
        ORDER BY times DESC
    """, (player_id,))
    role_stats = cursor.fetchall()

    # 计算胜率最高的角色
    best_role = None
    best_role_winrate = 0
    for rs in role_stats:
        wr = (rs["wins"] / rs["times"] * 100) if rs["times"] > 0 else 0
        if wr > best_role_winrate and rs["times"] >= 3:
            best_role = dict(rs)
            best_role["win_rate"] = round(wr, 1)
            best_role_winrate = wr

    conn.close()
    return render_template("player.html", request, {
        "player": player,
        "total_score": total_score, "total_matches": total_matches,
        "win_rate": win_rate, "avg_score": avg_score,
        "good_count": g[0], "good_winrate": g[1] or 0, "good_avg": g[2] or 0,
        "good_wins": round(g[0] * (g[1] or 0) / 100),
        "evil_count": e[0], "evil_winrate": e[1] or 0, "evil_avg": e[2] or 0,
        "evil_wins": round(e[0] * (e[1] or 0) / 100),
        "history": history_with_actions, "achievements": achievements,
        "role_stats": role_stats, "best_role": best_role,
        "admin": get_admin(request)
    })


@app.get("/admin", response_class=HTMLResponse)
async def admin_page(request: Request):
    admin = require_admin(request)
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM games ORDER BY id")
    games = cursor.fetchall()
    # 加载每个游戏的角色
    games_with_roles = []
    for g in games:
        cursor.execute("SELECT id, name, team, score_bonus, is_unique FROM roles WHERE game_id = ?", (g["id"],))
        roles = cursor.fetchall()
        gdict = dict(g)
        gdict["roles"] = [dict(r) for r in roles]
        games_with_roles.append(gdict)
    cursor.execute("SELECT id, name FROM players ORDER BY id")
    players = cursor.fetchall()
    cursor.execute("SELECT id, name, is_active FROM seasons ORDER BY started_at DESC")
    seasons = cursor.fetchall()
    cursor.execute("SELECT id, name, description, icon, condition_type, condition_value, is_preset FROM achievements ORDER BY id")
    achievements = cursor.fetchall()
    cursor.execute("""
        SELECT m.id, g.name, m.winner, m.played_at, COALESCE(m.notes, '') as notes,
               GROUP_CONCAT(p.name || '(' || CASE WHEN mp.team='good' THEN '好' ELSE '坏' END || ')', '、') as players_info
        FROM matches m JOIN games g ON m.game_id = g.id
        LEFT JOIN match_players mp ON m.id = mp.match_id
        LEFT JOIN players p ON mp.player_id = p.id
        GROUP BY m.id ORDER BY m.played_at DESC LIMIT 50
    """)
    matches = cursor.fetchall()
    cursor.execute("SELECT id, name, score_bonus, is_preset, game_id FROM game_actions ORDER BY game_id")
    actions = cursor.fetchall()
    cursor.execute("SELECT key, value FROM settings")
    settings_rows = cursor.fetchall()
    # 游戏名称映射（用于成就显示）
    game_names = {g["id"]: g["name"] for g in games}
    cursor.execute("SELECT id, name FROM roles")
    role_names = {r["id"]: r["name"] for r in cursor.fetchall()}
    conn.close()
    return render_template("admin.html", request, {
        "games": games_with_roles, "players": players,
        "matches": matches, "seasons": seasons, "achievements": achievements,
        "actions": actions, "admin": admin,
        "settings": {r["key"]: r["value"] for r in settings_rows},
        "game_names": game_names, "role_names": role_names
    })


@app.get("/admin/login", response_class=HTMLResponse)
async def login_page(request: Request):
    return render_template("login.html", request)


# ==================== 认证 API ====================

@app.post("/api/auth/login")
async def api_login(request: Request, username: str = Form(...), password: str = Form(...)):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT username FROM admins WHERE username = ? AND password_hash = ?",
                   (username, hash_password(password)))
    row = cursor.fetchone()
    conn.close()
    if row:
        request.session["admin"] = row[0]
        return RedirectResponse("/admin", status_code=302)
    return render_template("login.html", request, {"error": "用户名或密码错误"})


@app.get("/api/auth/logout")
async def api_logout(request: Request):
    request.session.clear()
    return RedirectResponse("/", status_code=302)


# ==================== 对局 API（含ELO） ====================

@app.get("/api/games/{game_id}/rules")
async def get_game_rules(game_id: int):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM games WHERE id = ?", (game_id,))
    game = cursor.fetchone()
    cursor.execute("SELECT id, name, team, score_bonus, is_unique FROM roles WHERE game_id = ?", (game_id,))
    roles = [dict(r) for r in cursor.fetchall()]
    conn.close()
    if not game:
        raise HTTPException(status_code=404)
    result = dict(game)
    result["roles"] = roles
    return result


@app.get("/api/games/{game_id}/role-presets")
async def get_role_presets(game_id: int, player_count: int = None):
    """获取基于人数的角色推荐配置"""
    conn = get_db()
    cursor = conn.cursor()
    if player_count:
        cursor.execute(
            "SELECT * FROM role_presets WHERE game_id = ? AND player_count = ? LIMIT 1",
            (game_id, player_count)
        )
    else:
        cursor.execute(
            "SELECT * FROM role_presets WHERE game_id = ? ORDER BY player_count",
            (game_id,)
        )
    rows = cursor.fetchall()
    conn.close()
    import json
    results = []
    for r in rows:
        d = dict(r)
        d["config"] = json.loads(d["config"])
        results.append(d)
    return results


@app.post("/api/matches")
async def create_match(request: Request):
    data = await request.json()
    game_id = data["game_id"]
    winner = data["winner"]
    players_data = data["players"]
    actions_data = data.get("actions", [])

    if len(players_data) < 2:
        return JSONResponse({"error": "至少需要2名玩家"}, status_code=400)

    has_good = any(p["team"] == "good" for p in players_data)
    has_evil = any(p["team"] == "evil" for p in players_data)
    if not has_good or not has_evil:
        return JSONResponse({"error": "需要好人和坏人阵营都有人"}, status_code=400)

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM games WHERE id = ?", (game_id,))
    game = cursor.fetchone()
    if not game:
        conn.close()
        return JSONResponse({"error": "游戏不存在"}, status_code=404)

    season = get_active_season(conn)
    season_id = season[0] if season else None
    use_elo = game["use_elo"]

    # 计算 ELO
    if use_elo:
        good_elos, evil_elos = [], []
        for p in players_data:
            cursor.execute("SELECT elo_rating FROM players WHERE id = ?", (p["player_id"],))
            el = cursor.fetchone()
            if el:
                (evil_elos if p["team"] == "evil" else good_elos).append(el[0])
        good_avg = sum(good_elos) / len(good_elos) if good_elos else 1000
        evil_avg = sum(evil_elos) / len(evil_elos) if evil_elos else 1000

    cursor.execute("INSERT INTO matches (game_id, winner, season_id) VALUES (?, ?, ?)",
                   (game_id, winner, season_id))
    match_id = cursor.lastrowid

    # 保存对局笔记
    notes = data.get("notes", "").strip()
    if notes:
        cursor.execute("UPDATE matches SET notes = ? WHERE id = ?", (notes, match_id))

    score_changes = {}
    elo_changes = {}
    pid_team_map = {}  # player_id -> team，用于操作加分时判断输赢
    lose_mode = game.get("lose_mode", "keep_bonus")

    for p in players_data:
        pid, team = p["player_id"], p["team"]
        pid_team_map[pid] = team
        role_id = p.get("role_id")

        # 自定义角色：当场创建
        if p.get("custom_role_name"):
            cr_name = p["custom_role_name"].strip()
            cr_bonus = int(p.get("custom_role_bonus") or 0)
            if cr_name:
                cursor.execute("SELECT id FROM roles WHERE game_id = ? AND name = ?", (game_id, cr_name))
                exist = cursor.fetchone()
                if exist:
                    role_id = exist[0]
                else:
                    cursor.execute(
                        "INSERT INTO roles (game_id, name, team, score_bonus) VALUES (?, ?, ?, ?)",
                        (game_id, cr_name, team, cr_bonus)
                    )
                    role_id = cursor.lastrowid

        # 基础得分
        if not use_elo:
            if team == winner:
                base_score = game["good_win_score"] if team == "good" else game["evil_win_score"]
            else:
                if lose_mode == "zero":
                    base_score = 0
                else:
                    base_score = -game["lose_penalty"]
            score = base_score
            elo_change = 0
        else:
            cursor.execute("SELECT elo_rating FROM players WHERE id = ?", (pid,))
            player_elo = cursor.fetchone()[0]
            opp_avg = evil_avg if team == "good" else good_avg
            actual = 1.0 if team == winner else 0.0
            new_elo = calc_elo(player_elo, opp_avg, actual)
            elo_change = new_elo - player_elo
            score = elo_change
            cursor.execute("UPDATE players SET elo_rating = ? WHERE id = ?", (new_elo, pid))

        # 角色加分
        role_bonus = 0
        if role_id:
            cursor.execute("SELECT score_bonus FROM roles WHERE id = ?", (role_id,))
            rb = cursor.fetchone()
            if rb:
                role_bonus = rb[0]
                if not use_elo and team != winner and lose_mode == "penalize_all":
                    role_bonus = -role_bonus  # 失败时角色加分变扣分
                elif not use_elo and team != winner and lose_mode == "zero":
                    role_bonus = 0  # 失败全0，不加角色分
                if role_bonus != 0:
                    score += role_bonus

        # 操作加分（先插入 match_player 才能关联 action）
        cursor.execute(
            "INSERT INTO match_players (match_id, player_id, team, role_id, score_change, elo_change) VALUES (?, ?, ?, ?, ?, ?)",
            (match_id, pid, team, role_id, score, elo_change)
        )
        score_changes[pid] = score
        elo_changes[pid] = elo_change

    # 处理操作加分（含自定义操作当场创建）
    action_scores = {}
    for act in actions_data:
        aid = act.get("action_id")
        pid = act["player_id"]

        # 自定义操作：当场创建
        if act.get("custom_action_name"):
            ca_name = act["custom_action_name"].strip()
            ca_bonus = int(act.get("custom_action_bonus") or 0)
            if ca_name:
                cursor.execute("SELECT id FROM game_actions WHERE game_id = ? AND name = ?", (game_id, ca_name))
                exist = cursor.fetchone()
                if exist:
                    aid = exist[0]
                else:
                    cursor.execute(
                        "INSERT INTO game_actions (game_id, name, score_bonus, is_preset) VALUES (?, ?, ?, 0)",
                        (game_id, ca_name, ca_bonus)
                    )
                    aid = cursor.lastrowid

        if not aid:
            continue
        cursor.execute("SELECT name, score_bonus FROM game_actions WHERE id = ?", (aid,))
        ab = cursor.fetchone()
        if ab:
            action_bonus = ab[1]
            # 根据失败规则调整操作加分
            if not use_elo:
                player_team = pid_team_map.get(pid)
                if player_team and player_team != winner:
                    if lose_mode == "zero":
                        action_bonus = 0
                    elif lose_mode == "penalize_all":
                        action_bonus = -action_bonus
            cursor.execute(
                "INSERT INTO match_player_actions (match_id, player_id, action_id) VALUES (?, ?, ?)",
                (match_id, pid, aid)
            )
            if action_bonus != 0:
                cursor.execute(
                    "UPDATE match_players SET score_change = score_change + ? WHERE match_id = ? AND player_id = ?",
                    (action_bonus, match_id, pid)
                )
                score_changes[pid] = score_changes.get(pid, 0) + action_bonus
            action_scores[pid] = action_bonus

    conn.commit()

    all_new_achievements = {}
    if get_setting(conn, "enable_achievements", "1") == "1":
        for p in players_data:
            new = check_achievements(conn, p["player_id"])
            if new:
                all_new_achievements[p["player_id"]] = new

    auto_bak = get_setting(conn, "auto_backup", "0") == "1"
    conn.close()

    # 开启自动备份时，对局提交后立即备份
    if auto_bak:
        auto_backup_db()

    return {"success": True, "match_id": match_id, "scores": score_changes,
            "elo_changes": elo_changes, "action_scores": action_scores,
            "new_achievements": all_new_achievements, "elo_mode": bool(use_elo)}


@app.delete("/api/matches/{match_id}")
async def delete_match(request: Request, match_id: int):
    require_admin(request)
    conn = get_db()
    cursor = conn.cursor()

    # 回滚 ELO
    cursor.execute("SELECT player_id, elo_change FROM match_players WHERE match_id = ? AND elo_change != 0", (match_id,))
    for row in cursor.fetchall():
        cursor.execute("UPDATE players SET elo_rating = elo_rating - ? WHERE id = ?", (row[1], row[0]))

    cursor.execute("DELETE FROM matches WHERE id = ?", (match_id,))
    conn.commit()
    conn.close()
    return {"success": True}


@app.post("/api/matches/batch-delete")
async def batch_delete_matches(request: Request):
    """批量删除对局"""
    require_admin(request)
    data = await request.json()
    ids = data.get("ids", [])
    if not ids:
        return JSONResponse({"error": "请选择至少一条对局"}, status_code=400)

    conn = get_db()
    cursor = conn.cursor()
    deleted = 0
    for mid in ids:
        try:
            cursor.execute("SELECT player_id, elo_change FROM match_players WHERE match_id = ? AND elo_change != 0", (mid,))
            for row in cursor.fetchall():
                cursor.execute("UPDATE players SET elo_rating = elo_rating - ? WHERE id = ?", (row[1], row[0]))
            cursor.execute("DELETE FROM matches WHERE id = ?", (mid,))
            deleted += 1
        except:
            pass
    conn.commit()
    conn.close()
    return {"success": True, "deleted": deleted}


@app.get("/api/matches/history")
async def match_history(game_id: Optional[int] = None, limit: int = 30):
    conn = get_db()
    cursor = conn.cursor()
    base = """
        SELECT m.id, g.name as game_name, m.winner, m.played_at, COALESCE(m.notes,'') as notes,
               GROUP_CONCAT(p.name || ':' || mp.team || ':' || mp.score_change || ':' || COALESCE(r.name,'') || ':' || COALESCE(r.score_bonus,0), '、') as detail
        FROM matches m JOIN games g ON m.game_id = g.id
        LEFT JOIN match_players mp ON m.id = mp.match_id
        LEFT JOIN players p ON mp.player_id = p.id
        LEFT JOIN roles r ON mp.role_id = r.id
    """
    if game_id:
        cursor.execute(base + " WHERE m.game_id = ? GROUP BY m.id ORDER BY m.played_at DESC LIMIT ?", (game_id, limit))
    else:
        cursor.execute(base + " GROUP BY m.id ORDER BY m.played_at DESC LIMIT ?", (limit,))
    results = []
    for r in cursor.fetchall():
        d = dict(r)
        # 查询每局所有操作
        cursor.execute("""
            SELECT mpa.player_id, p.name, ga.name as action_name, ga.score_bonus
            FROM match_player_actions mpa
            JOIN game_actions ga ON mpa.action_id = ga.id
            JOIN players p ON mpa.player_id = p.id
            WHERE mpa.match_id = ?
        """, (d["id"],))
        actions = [dict(a) for a in cursor.fetchall()]
        d["actions"] = actions
        results.append(d)
    return results


@app.get("/api/matches/{match_id}")
async def match_detail(match_id: int):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT mp.id as mp_id, mp.player_id, p.name, mp.team, mp.score_change, mp.elo_change,
               COALESCE(r.name, '') as role_name, COALESCE(r.score_bonus, 0) as role_bonus
        FROM match_players mp
        JOIN players p ON mp.player_id = p.id
        LEFT JOIN roles r ON mp.role_id = r.id
        WHERE mp.match_id = ?
    """, (match_id,))
    players_raw = cursor.fetchall()

    players = []
    for pr in players_raw:
        pd = dict(pr)
        # 查询该玩家本局的操作加分
        cursor.execute("""
            SELECT ga.name, ga.score_bonus
            FROM match_player_actions mpa
            JOIN game_actions ga ON mpa.action_id = ga.id
            WHERE mpa.match_id = ? AND mpa.player_id = ?
        """, (match_id, pr["player_id"]))
        actions = [dict(a) for a in cursor.fetchall()]
        pd["actions"] = actions
        pd["action_total"] = sum(a["score_bonus"] for a in actions)
        pd["base_score"] = pd["score_change"] - pd["role_bonus"] - pd["action_total"]
        players.append(pd)
    conn.close()
    return players


@app.get("/api/balance/suggest")
async def balance_suggest(game_id: int = None):
    """根据ELO给出平衡建议"""
    conn = get_db()
    cursor = conn.cursor()
    if game_id:
        cursor.execute("SELECT id, name, elo_rating FROM players WHERE id IN (SELECT DISTINCT player_id FROM match_players mp JOIN matches m ON mp.match_id = m.id WHERE m.game_id = ?) ORDER BY elo_rating DESC", (game_id,))
    else:
        cursor.execute("SELECT id, name, elo_rating FROM players ORDER BY elo_rating DESC")
    players = [dict(r) for r in cursor.fetchall()]
    conn.close()

    if len(players) < 2:
        return {"players": players, "suggestions": []}

    # 贪心分配：按 ELO 排序，轮流分配到两队
    team_a, team_b = [], []
    sum_a, sum_b = 0, 0
    for p in players:
        if sum_a <= sum_b:
            team_a.append(p)
            sum_a += p["elo_rating"]
        else:
            team_b.append(p)
            sum_b += p["elo_rating"]

    suggestions = []
    diff = abs(sum_a - sum_b)
    if diff > 100:
        suggestions.append(f"两队ELO差距 {diff} 较大，建议将 {team_a[-1]['name']} 和 {team_b[1]['name'] if len(team_b)>1 else team_b[0]['name']} 交换以平衡")

    return {
        "players": players,
        "team_a": team_a, "team_b": team_b,
        "team_a_avg": round(sum_a / len(team_a)) if team_a else 0,
        "team_b_avg": round(sum_b / len(team_b)) if team_b else 0,
        "suggestions": suggestions
    }


# ==================== 赛季 API ====================

@app.post("/api/seasons")
async def create_season(request: Request):
    require_admin(request)
    data = await request.json()
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE seasons SET is_active = 0 WHERE is_active = 1")
    cursor.execute("INSERT INTO seasons (name, is_active) VALUES (?, 1)", (data["name"],))
    sid = cursor.lastrowid
    conn.commit()
    conn.close()
    return {"success": True, "id": sid}


@app.put("/api/seasons/{season_id}/activate")
async def activate_season(request: Request, season_id: int):
    require_admin(request)
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE seasons SET is_active = 0")
    cursor.execute("UPDATE seasons SET is_active = 1, ended_at = NULL WHERE id = ?", (season_id,))
    conn.commit()
    conn.close()
    return {"success": True}


@app.put("/api/seasons/{season_id}/end")
async def end_season(request: Request, season_id: int):
    require_admin(request)
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE seasons SET is_active = 0, ended_at = CURRENT_TIMESTAMP WHERE id = ?", (season_id,))
    conn.commit()
    conn.close()
    return {"success": True}


# ==================== 成就 API ====================

@app.post("/api/achievements")
async def create_achievement(request: Request):
    require_admin(request)
    data = await request.json()
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO achievements (name, description, icon, condition_type, condition_value, game_id, role_id, is_preset) VALUES (?, ?, ?, ?, ?, ?, ?, 0)",
        (data["name"], data.get("description", ""), data.get("icon", "🏅"),
         data["condition_type"], data["condition_value"],
         data.get("game_id"), data.get("role_id"))
    )
    conn.commit()
    conn.close()
    return {"success": True}


@app.delete("/api/achievements/{ach_id}")
async def delete_achievement(request: Request, ach_id: int):
    require_admin(request)
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT is_preset FROM achievements WHERE id = ?", (ach_id,))
    row = cursor.fetchone()
    if row and row[0]:
        conn.close()
        return JSONResponse({"error": "不能删除预设成就"}, status_code=400)
    cursor.execute("DELETE FROM achievements WHERE id = ?", (ach_id,))
    conn.commit()
    conn.close()
    return {"success": True}


@app.get("/api/achievements/progress/{player_id}")
async def achievement_progress(player_id: int):
    """获取玩家所有成就的当前进度"""
    conn = get_db()
    cursor = conn.cursor()

    # 已获得的成就
    cursor.execute("SELECT achievement_id FROM player_achievements WHERE player_id = ?", (player_id,))
    earned = {r[0] for r in cursor.fetchall()}

    # 计算该玩家的各项统计值
    cursor.execute("SELECT COUNT(*) FROM match_players mp JOIN matches m ON mp.match_id = m.id WHERE mp.player_id = ?", (player_id,))
    total_matches = cursor.fetchone()[0]

    cursor.execute("SELECT SUM(CASE WHEN mp.team = m.winner THEN 1 ELSE 0 END) FROM match_players mp JOIN matches m ON mp.match_id = m.id WHERE mp.player_id = ?", (player_id,))
    total_wins = cursor.fetchone()[0] or 0

    cursor.execute("SELECT COALESCE(SUM(mp.score_change), 0) FROM match_players mp WHERE mp.player_id = ?", (player_id,))
    total_score = cursor.fetchone()[0]

    cursor.execute("SELECT COALESCE(MAX(mp.score_change), 0) FROM match_players mp WHERE mp.player_id = ?", (player_id,))
    high_score = cursor.fetchone()[0]

    cursor.execute("""SELECT COUNT(*) FROM match_players mp
        WHERE mp.player_id = ? AND mp.score_change > 0
          AND mp.score_change = (SELECT MAX(mp2.score_change) FROM match_players mp2 WHERE mp2.match_id = mp.match_id)""", (player_id,))
    mvp_count = cursor.fetchone()[0]

    cursor.execute("SELECT mp.team = m.winner as is_win FROM match_players mp JOIN matches m ON mp.match_id = m.id WHERE mp.player_id = ? ORDER BY m.played_at DESC", (player_id,))
    results = cursor.fetchall()
    cw, cl = 0, 0
    for r in results:
        if r[0]: cw += 1
        else: break
    for r in results:
        if not r[0]: cl += 1
        else: break

    cursor.execute("""SELECT COUNT(*), ROUND(CAST(SUM(CASE WHEN mp.team = m.winner THEN 1 ELSE 0 END) AS FLOAT)/NULLIF(COUNT(*),0)*100)
        FROM match_players mp JOIN matches m ON mp.match_id = m.id WHERE mp.player_id = ? AND mp.team='good'""", (player_id,))
    g = cursor.fetchone()
    cursor.execute("""SELECT COUNT(*), ROUND(CAST(SUM(CASE WHEN mp.team = m.winner THEN 1 ELSE 0 END) AS FLOAT)/NULLIF(COUNT(*),0)*100)
        FROM match_players mp JOIN matches m ON mp.match_id = m.id WHERE mp.player_id = ? AND mp.team='evil'""", (player_id,))
    e = cursor.fetchone()
    cursor.execute("SELECT elo_rating FROM players WHERE id = ?", (player_id,))
    elo = cursor.fetchone()[0]

    cursor.execute("SELECT m.game_id, COUNT(*), SUM(CASE WHEN mp.team=m.winner THEN 1 ELSE 0 END) FROM match_players mp JOIN matches m ON mp.match_id=m.id WHERE mp.player_id=? GROUP BY m.game_id", (player_id,))
    game_stats = {row[0]: {"matches": row[1], "wins": row[2]} for row in cursor.fetchall()}

    cursor.execute("SELECT mp.role_id, COUNT(*), SUM(CASE WHEN mp.team=m.winner THEN 1 ELSE 0 END) FROM match_players mp JOIN matches m ON mp.match_id=m.id WHERE mp.player_id=? AND mp.role_id IS NOT NULL GROUP BY mp.role_id", (player_id,))
    role_stats = {row[0]: {"matches": row[1], "wins": row[2]} for row in cursor.fetchall()}

    stats = {
        "total_matches": total_matches, "total_wins": total_wins, "total_score": total_score,
        "high_score": high_score, "mvp_count": mvp_count, "consecutive_wins": cw,
        "consecutive_loses": cl, "good_win_rate": g[1] or 0, "evil_win_rate": e[1] or 0,
        "elo_rating": elo,
    }

    cursor.execute("SELECT id, name, description, icon, condition_type, condition_value, game_id, role_id, is_preset FROM achievements")
    achievements = []
    for ach in cursor.fetchall():
        aid, name, desc, icon, ctype, cval, gid, rid, isp = ach
        current = None
        if ctype in stats:
            current = stats[ctype]
        elif ctype == "game_matches":
            current = game_stats[gid]["matches"] if gid in game_stats else 0
        elif ctype == "game_wins":
            current = game_stats[gid]["wins"] if gid in game_stats else 0
        elif ctype == "role_matches":
            current = role_stats[rid]["matches"] if rid in role_stats else 0
        elif ctype == "role_wins":
            current = role_stats[rid]["wins"] if rid in role_stats else 0

        achievements.append({
            "id": aid, "name": name, "description": desc, "icon": icon,
            "condition_type": ctype, "condition_value": cval,
            "current_value": current if current is not None else 0,
            "earned": aid in earned,
            "game_id": gid, "role_id": rid, "is_preset": isp
        })

    # 补充游戏和角色名称
    cursor.execute("SELECT id, name FROM games")
    game_names = {r[0]: r[1] for r in cursor.fetchall()}
    cursor.execute("SELECT id, name FROM roles")
    role_names = {r[0]: r[1] for r in cursor.fetchall()}
    for a in achievements:
        if a["game_id"]:
            a["game_name"] = game_names.get(a["game_id"], "")
        if a["role_id"]:
            a["role_name"] = role_names.get(a["role_id"], "")

    conn.close()
    return achievements


# ==================== 备份 API ====================

@app.get("/api/backup/download")
async def download_backup(request: Request):
    require_admin(request)
    from starlette.responses import FileResponse
    file_path = __import__('database').DB_PATH
    return FileResponse(file_path,
        media_type="application/octet-stream",
        filename="game_backup.db")


@app.post("/api/backup/restore")
async def restore_backup(request: Request):
    require_admin(request)
    form = await request.form()
    file = form.get("file")
    if not file:
        return JSONResponse({"error": "请选择备份文件"}, status_code=400)

    content = await file.read()
    # 验证是否为有效 SQLite
    if len(content) < 100 or content[:16] != b'SQLite format 3\x00':
        return JSONResponse({"error": "无效的数据库文件"}, status_code=400)

    file_path = __import__('database').DB_PATH
    import shutil
    # 先备份当前
    shutil.copy2(file_path, file_path + ".auto_bak")
    with open(file_path, "wb") as f:
        f.write(content)
    return {"success": True, "message": "恢复成功，已备份旧文件为 game.db.auto_bak"}


@app.get("/api/backup/auto")
async def list_auto_backups(request: Request):
    """列出所有自动备份文件"""
    require_admin(request)
    ensure_backup_dir()
    files = sorted(glob.glob(os.path.join(BACKUP_DIR, "auto_backup_*.db")), reverse=True)
    backups = []
    for f in files:
        fname = os.path.basename(f)
        fsize = os.path.getsize(f)
        if fsize < 1024 * 1024:
            fsize_str = f"{fsize / 1024:.1f} KB"
        else:
            fsize_str = f"{fsize / 1024 / 1024:.1f} MB"
        mtime = datetime.fromtimestamp(os.path.getmtime(f)).strftime("%Y-%m-%d %H:%M:%S")
        backups.append({"filename": fname, "size": fsize_str, "time": mtime})
    return backups


@app.get("/api/backup/auto/{filename}")
async def download_auto_backup(request: Request, filename: str):
    """下载某个自动备份文件"""
    require_admin(request)
    from starlette.responses import FileResponse
    ensure_backup_dir()
    file_path = os.path.join(BACKUP_DIR, filename)
    if not os.path.exists(file_path):
        return JSONResponse({"error": "备份文件不存在"}, status_code=404)
    return FileResponse(file_path, media_type="application/octet-stream", filename=filename)


@app.delete("/api/backup/auto/{filename}")
async def delete_auto_backup(request: Request, filename: str):
    """删除某个自动备份文件"""
    require_admin(request)
    ensure_backup_dir()
    file_path = os.path.join(BACKUP_DIR, filename)
    if not os.path.exists(file_path):
        return JSONResponse({"error": "备份文件不存在"}, status_code=404)
    os.remove(file_path)
    return {"success": True}


# ==================== 角色 API ====================

@app.post("/api/roles")
async def create_role(request: Request):
    require_admin(request)
    data = await request.json()
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO roles (game_id, name, team, score_bonus) VALUES (?, ?, ?, ?)",
        (data["game_id"], data["name"], data["team"], data.get("score_bonus", 0))
    )
    rid = cursor.lastrowid
    conn.commit()
    conn.close()
    return {"success": True, "id": rid}


@app.delete("/api/roles/{role_id}")
async def delete_role(request: Request, role_id: int):
    require_admin(request)
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM roles WHERE id = ?", (role_id,))
    conn.commit()
    conn.close()
    return {"success": True}


# ==================== 导出 API ====================

def _ts_filename(base_name):
    """给文件名加上时间后缀，避免重名覆盖"""
    name, ext = os.path.splitext(base_name)
    ts = time.strftime("%Y%m%d_%H%M%S")
    return f"{name}_{ts}{ext}"

def _t(val):
    """将英文 winner/team 翻译成中文"""
    if val == 'good': return '好人'
    if val == 'evil': return '坏人'
    if val == 'draw': return '平局'
    return str(val) if val else ''


@app.get("/api/export/rankings")
async def export_rankings(request: Request, season_id: str = ""):
    sid_raw = int(season_id) if season_id else None
    conn = get_db()
    cursor = conn.cursor()

    if sid_raw:
        # 按赛季筛选：该赛季积分 + 全部数据汇总
        cursor.execute("""
            SELECT p.name, p.elo_rating,
                   COALESCE(SUM(CASE WHEN m.season_id = ? THEN mp.score_change ELSE 0 END), 0) as season_score,
                   COUNT(CASE WHEN m.season_id = ? THEN mp.id END) as season_matches,
                   COALESCE(SUM(mp.score_change), 0) as total_score,
                   COUNT(mp.id) as total_matches,
                   SUM(CASE WHEN mp.team = m.winner THEN 1 ELSE 0 END) as total_wins,
                   COUNT(mp.id) - SUM(CASE WHEN mp.team = m.winner THEN 1 ELSE 0 END) as total_losses,
                   ROUND(CAST(SUM(CASE WHEN mp.team = m.winner THEN 1 ELSE 0 END) AS FLOAT) / NULLIF(COUNT(mp.id), 0) * 100, 1) as win_rate,
                   SUM(CASE WHEN mp.team = 'good' THEN 1 ELSE 0 END) as good_count,
                   SUM(CASE WHEN mp.team = 'evil' THEN 1 ELSE 0 END) as evil_count
            FROM players p
            LEFT JOIN match_players mp ON p.id = mp.player_id
            LEFT JOIN matches m ON mp.match_id = m.id
            GROUP BY p.id
            HAVING COUNT(CASE WHEN m.season_id = ? THEN mp.id END) > 0
            ORDER BY season_score DESC
        """, (sid_raw, sid_raw, sid_raw))
    else:
        cursor.execute("""
            SELECT p.name, p.elo_rating, 0 as season_score, 0 as season_matches,
                   COALESCE(SUM(mp.score_change), 0) as total_score,
                   COUNT(mp.id) as total_matches,
                   SUM(CASE WHEN mp.team = m.winner THEN 1 ELSE 0 END) as total_wins,
                   COUNT(mp.id) - SUM(CASE WHEN mp.team = m.winner THEN 1 ELSE 0 END) as total_losses,
                   ROUND(CAST(SUM(CASE WHEN mp.team = m.winner THEN 1 ELSE 0 END) AS FLOAT) / NULLIF(COUNT(mp.id), 0) * 100, 1) as win_rate,
                   SUM(CASE WHEN mp.team = 'good' THEN 1 ELSE 0 END) as good_count,
                   SUM(CASE WHEN mp.team = 'evil' THEN 1 ELSE 0 END) as evil_count
            FROM players p
            LEFT JOIN match_players mp ON p.id = mp.player_id
            LEFT JOIN matches m ON mp.match_id = m.id
            GROUP BY p.id HAVING COUNT(mp.id) > 0
            ORDER BY total_score DESC
        """)

    rows = cursor.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "排行榜"

    headers = ["排名", "玩家", "ELO评分", "总积分", "对局数", "胜场", "负场", "胜率", "好人局", "坏人局"]
    col_widths = [8, 14, 12, 10, 10, 10, 10, 10, 10, 10]
    for i, r in enumerate(rows, 1):
        ws.append([i, r["name"], r["elo_rating"], r["total_score"], r["total_matches"],
                   r["total_wins"], r["total_losses"], f"{r['win_rate']}%",
                   r["good_count"], r["evil_count"]])

    style_excel_sheet(ws, headers, col_widths)
    return make_excel_response(wb, _ts_filename("rankings.xlsx"))


@app.get("/api/export/matches")
async def export_matches(request: Request, game_id: int = None):
    conn = get_db()
    cursor = conn.cursor()

    # ========== Sheet1: 对局摘要（增加赛季+备注） ==========
    if game_id:
        cursor.execute("""
            SELECT m.id, g.name as game, COALESCE(s.name, '-') as season,
                   m.winner, m.played_at, m.notes
            FROM matches m
            JOIN games g ON m.game_id = g.id
            LEFT JOIN seasons s ON m.season_id = s.id
            WHERE m.game_id = ? ORDER BY m.played_at DESC
        """, (game_id,))
    else:
        cursor.execute("""
            SELECT m.id, g.name as game, COALESCE(s.name, '-') as season,
                   m.winner, m.played_at, m.notes
            FROM matches m
            JOIN games g ON m.game_id = g.id
            LEFT JOIN seasons s ON m.season_id = s.id
            ORDER BY m.played_at DESC
        """)
    match_rows = cursor.fetchall()

    # ========== Sheet2: 对局详情（展开每人一行，增加角色信息） ==========
    if match_rows:
        match_ids = [m["id"] for m in match_rows]
        placeholders = ",".join("?" for _ in match_ids)
        if game_id:
            cursor.execute(f"""
                SELECT mp.match_id, p.name, mp.team,
                       COALESCE(r.name, '-') as role,
                       mp.score_change, mp.elo_change
                FROM match_players mp
                JOIN players p ON mp.player_id = p.id
                LEFT JOIN roles r ON mp.role_id = r.id
                JOIN matches m ON mp.match_id = m.id
                WHERE mp.match_id IN ({placeholders}) AND m.game_id = ?
                ORDER BY mp.match_id, mp.team
            """, (*match_ids, game_id))
        else:
            cursor.execute(f"""
                SELECT mp.match_id, p.name, mp.team,
                       COALESCE(r.name, '-') as role,
                       mp.score_change, mp.elo_change
                FROM match_players mp
                JOIN players p ON mp.player_id = p.id
                LEFT JOIN roles r ON mp.role_id = r.id
                WHERE mp.match_id IN ({placeholders})
                ORDER BY mp.match_id, mp.team
            """, match_ids)
        detail_rows = cursor.fetchall()
    else:
        detail_rows = []

    conn.close()

    wb = Workbook()
    wb.remove(wb.active)

    # ========== Sheet1: 对局摘要 ==========
    ws1 = wb.create_sheet("对局摘要", 0)
    headers1 = ["对局ID", "游戏", "赛季", "获胜方", "时间", "备注"]
    col_widths1 = [10, 14, 12, 10, 20, 20]
    for r in match_rows:
        ws1.append([r["id"], r["game"], r["season"], _t(r["winner"]),
                    r["played_at"], r["notes"] or ""])
    style_excel_sheet(ws1, headers1, col_widths1)

    # ========== Sheet2: 对局详情（每人一行） ==========
    ws2 = wb.create_sheet("对局详情", 1)
    headers2 = ["对局ID", "玩家", "阵营", "角色", "得分变化", "ELO变化"]
    col_widths2 = [10, 14, 10, 12, 10, 10]
    for r in detail_rows:
        ws2.append([r["match_id"], r["name"], _t(r["team"]),
                    r["role"], r["score_change"], r["elo_change"]])
    style_excel_sheet(ws2, headers2, col_widths2)

    return make_excel_response(wb, _ts_filename("matches.xlsx"))


@app.get("/api/export/player/{player_id}")
async def export_player(request: Request, player_id: int):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT g.name as game, COALESCE(s.name, '-') as season,
               m.winner, mp.team, COALESCE(r.name, '-') as role,
               mp.score_change, mp.elo_change, m.played_at
        FROM match_players mp
        JOIN matches m ON mp.match_id = m.id
        JOIN games g ON m.game_id = g.id
        LEFT JOIN seasons s ON m.season_id = s.id
        LEFT JOIN roles r ON mp.role_id = r.id
        WHERE mp.player_id = ? ORDER BY m.played_at DESC
    """, (player_id,))
    rows = cursor.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "玩家对局"

    headers = ["游戏", "赛季", "获胜方", "玩家阵营", "角色", "得分变化", "ELO变化", "时间"]
    col_widths = [14, 10, 10, 10, 12, 10, 10, 20]
    for r in rows:
        ws.append([r["game"], r["season"], _t(r["winner"]), _t(r["team"]),
                   r["role"], r["score_change"], r["elo_change"], r["played_at"]])

    style_excel_sheet(ws, headers, col_widths)
    return make_excel_response(wb, _ts_filename(f"player_{player_id}.xlsx"))


# ==================== 战报 API ====================

@app.get("/api/report/{match_id}")
async def match_report(match_id: int):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT m.*, g.name as game_name FROM matches m JOIN games g ON m.game_id = g.id WHERE m.id = ?", (match_id,))
    match = cursor.fetchone()
    if not match:
        conn.close()
        return JSONResponse({"error": "对局不存在"}, status_code=404)

    cursor.execute("""
        SELECT mp.team, mp.score_change, mp.elo_change, p.name, p.elo_rating,
               COALESCE(r.name, '') as role_name
        FROM match_players mp
        JOIN players p ON mp.player_id = p.id
        LEFT JOIN roles r ON mp.role_id = r.id
        WHERE mp.match_id = ?
    """, (match_id,))
    players = [dict(r) for r in cursor.fetchall()]
    conn.close()

    winner_label = "好人阵营" if match["winner"] == "good" else "坏人阵营"
    good_players = [p for p in players if p["team"] == "good"]
    evil_players = [p for p in players if p["team"] == "evil"]

    report = {
        "title": f"🎮 {match['game_name']} 对局战报",
        "time": match["played_at"],
        "result": f"{winner_label} 获胜!",
        "good_team": good_players,
        "evil_team": evil_players,
        "summary": f"本局{len(players)}人参与，最终{winner_label}获胜"
    }

    mvp = max(players, key=lambda x: x["score_change"])
    report["mvp"] = {"name": mvp["name"], "score": mvp["score_change"]}
    report["summary"] += f"，MVP: {mvp['name']}({mvp['score_change']:+d}分)"

    return report


# ==================== 玩家 API ====================

@app.get("/api/players")
async def get_players():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, name, avatar, elo_rating FROM players ORDER BY id")
    return [dict(r) for r in cursor.fetchall()]


@app.post("/api/players")
async def create_player(request: Request):
    data = await request.json()
    name = data["name"].strip()
    if not name:
        return JSONResponse({"error": "名字不能为空"}, status_code=400)
    conn = get_db()
    cursor = conn.cursor()
    # 已存在则返回已有记录
    cursor.execute("SELECT id, name, avatar, elo_rating FROM players WHERE name = ?", (name,))
    exist = cursor.fetchone()
    if exist:
        conn.close()
        return {"success": True, "id": exist["id"], "name": exist["name"], "existed": True}
    cursor.execute("INSERT INTO players (name) VALUES (?)", (name,))
    conn.commit()
    pid = cursor.lastrowid
    conn.close()
    return {"success": True, "id": pid, "name": name, "existed": False}


@app.delete("/api/players/{player_id}")
async def delete_player(request: Request, player_id: int):
    require_admin(request)
    conn = get_db()
    cursor = conn.cursor()
    try:
        # 清理关联的外键数据（match_player_actions → match_players → player_achievements → players）
        cursor.execute("""
            DELETE FROM match_player_actions WHERE player_id = ?
        """, (player_id,))
        cursor.execute("""
            DELETE FROM match_player_actions 
            WHERE match_id IN (SELECT match_id FROM match_players WHERE player_id = ?)
        """, (player_id,))
        cursor.execute("DELETE FROM match_players WHERE player_id = ?", (player_id,))
        cursor.execute("DELETE FROM player_achievements WHERE player_id = ?", (player_id,))
        cursor.execute("DELETE FROM players WHERE id = ?", (player_id,))
        conn.commit()
    except Exception as e:
        conn.close()
        return JSONResponse({"error": f"删除失败: {str(e)}"}, status_code=500)
    conn.close()
    return {"success": True}


# ==================== 游戏 API ====================

@app.get("/api/games")
async def get_games():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM games ORDER BY id")
    return [dict(r) for r in cursor.fetchall()]


@app.post("/api/games")
async def create_game(request: Request):
    require_admin(request)
    data = await request.json()
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "INSERT INTO games (name, good_win_score, evil_win_score, lose_penalty, lose_mode, use_elo) VALUES (?, ?, ?, ?, ?, ?)",
            (data["name"], data.get("good_win_score", 5), data.get("evil_win_score", 8),
             data.get("lose_penalty", 3), data.get("lose_mode", "keep_bonus"), data.get("use_elo", 0))
        )
        conn.commit()
        gid = cursor.lastrowid
        conn.close()
        return {"success": True, "id": gid}
    except Exception:
        conn.close()
        return JSONResponse({"error": "游戏名已存在"}, status_code=400)


@app.put("/api/games/{game_id}")
async def update_game(request: Request, game_id: int):
    require_admin(request)
    data = await request.json()
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE games SET name=?, good_win_score=?, evil_win_score=?, lose_penalty=?, lose_mode=?, use_elo=? WHERE id=?",
        (data["name"], data.get("good_win_score", 5), data.get("evil_win_score", 8),
         data.get("lose_penalty", 3), data.get("lose_mode", "keep_bonus"),
         data.get("use_elo", 0), game_id)
    )
    conn.commit()
    conn.close()
    return {"success": True}


# ==================== 预设阵容 API ====================

@app.get("/api/presets")
async def get_presets():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, name, game_id, config, created_at FROM game_presets ORDER BY created_at DESC")
    return [dict(r) for r in cursor.fetchall()]


@app.post("/api/presets")
async def create_preset(request: Request):
    data = await request.json()
    conn = get_db()
    cursor = conn.cursor()
    import json
    cursor.execute(
        "INSERT INTO game_presets (name, game_id, config) VALUES (?, ?, ?)",
        (data["name"], data["game_id"], json.dumps(data["config"], ensure_ascii=False))
    )
    conn.commit()
    pid = cursor.lastrowid
    conn.close()
    return {"success": True, "id": pid}


@app.delete("/api/presets/{preset_id}")
async def delete_preset(request: Request, preset_id: int):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM game_presets WHERE id = ?", (preset_id,))
    conn.commit()
    conn.close()
    return {"success": True}


# ==================== 操作加分 API ====================

@app.get("/api/games/{game_id}/actions")
async def get_game_actions(game_id: int):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, name, score_bonus, is_preset FROM game_actions WHERE game_id = ?", (game_id,))
    return [dict(r) for r in cursor.fetchall()]


@app.post("/api/actions")
async def create_action(request: Request):
    require_admin(request)
    data = await request.json()
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO game_actions (game_id, name, score_bonus) VALUES (?, ?, ?)",
        (data["game_id"], data["name"], data.get("score_bonus", 0))
    )
    aid = cursor.lastrowid
    conn.commit()
    conn.close()
    return {"success": True, "id": aid}


@app.delete("/api/actions/{action_id}")
async def delete_action(request: Request, action_id: int):
    require_admin(request)
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT is_preset FROM game_actions WHERE id = ?", (action_id,))
    row = cursor.fetchone()
    if row and row[0]:
        conn.close()
        return JSONResponse({"error": "不能删除预设操作"}, status_code=400)
    cursor.execute("DELETE FROM game_actions WHERE id = ?", (action_id,))
    conn.commit()
    conn.close()
    return {"success": True}


# ==================== 自定义角色 API ====================

@app.post("/api/roles/custom")
async def create_custom_role(request: Request):
    """记录对局时即时创建自定义角色，无需管理员权限"""
    data = await request.json()
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM roles WHERE game_id = ? AND name = ?", (data["game_id"], data["name"]))
    exist = cursor.fetchone()
    if exist:
        conn.close()
        return {"success": True, "id": exist[0], "existed": True}
    cursor.execute(
        "INSERT INTO roles (game_id, name, team, score_bonus) VALUES (?, ?, ?, ?)",
        (data["game_id"], data["name"], data.get("team", "good"), data.get("score_bonus", 0))
    )
    conn.commit()
    rid = cursor.lastrowid
    conn.close()
    return {"success": True, "id": rid, "existed": False}


# ==================== 设置 API ====================

@app.get("/api/settings")
async def get_settings():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT key, value FROM settings")
    rows = cursor.fetchall()
    conn.close()
    return {r["key"]: r["value"] for r in rows}


@app.put("/api/settings")
async def update_settings(request: Request):
    require_admin(request)
    data = await request.json()
    conn = get_db()
    cursor = conn.cursor()
    for key, value in data.items():
        cursor.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, value))
    conn.commit()
    conn.close()
    return {"success": True}


@app.put("/api/games/{game_id}/features")
async def update_game_features(request: Request, game_id: int):
    require_admin(request)
    data = await request.json()
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE games SET enable_roles=?, enable_actions=? WHERE id=?",
        (data.get("enable_roles", 1), data.get("enable_actions", 1), game_id)
    )
    conn.commit()
    conn.close()
    return {"success": True}


# ==================== EXE 启动入口 ====================
if __name__ == "__main__":
    _crash_log("init: entering startup block")

    import socket

    def safe_print(msg, fallback=""):
        try:
            print(msg, flush=True)
        except (UnicodeEncodeError, UnicodeDecodeError):
            if fallback:
                print(fallback, flush=True)

    def get_local_ip():
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            ip = s.getsockname()[0]
            s.close()
            return ip
        except Exception:
            return "127.0.0.1"

    exit_code = 0
    try:
        host = "0.0.0.0"
        port = 8000
        open_browser = True
        browser_path = None
        use_share = False

        i = 1
        while i < len(sys.argv):
            arg = sys.argv[i]
            if arg == "--local":
                host = "127.0.0.1"
            elif arg in ("-p", "--port") and i + 1 < len(sys.argv):
                i += 1
                port = int(sys.argv[i])
            elif arg in ("-b", "--browser") and i + 1 < len(sys.argv):
                i += 1
                browser_path = sys.argv[i]
            elif arg in ("--no-browser", "-nb"):
                open_browser = False
            elif arg in ("--share", "-s"):
                use_share = True
            elif arg in ("-h", "--help"):
                safe_print("Usage: 桌游排行.exe [options]")
                safe_print("")
                safe_print("Options:")
                safe_print("  -p, --port PORT      HTTP 端口 (默认 8000)")
                safe_print("  -b, --browser PATH   指定浏览器路径")
                safe_print("  --local              仅监听 127.0.0.1")
                safe_print("  --no-browser, -nb    不自动打开浏览器")
                safe_print("  --share, -s           开启公网隧道 (ngrok)")
                safe_print("  -h, --help           显示此帮助")
                safe_print("")
                sys.exit(0)
            i += 1

        local_ip = get_local_ip()

        safe_print("=" * 50)
        safe_print("  Board Game Ranking System")
        safe_print("=" * 50)
        safe_print(f"  Local:    http://127.0.0.1:{port}")
        safe_print(f"  Network:  http://{local_ip}:{port}")
        safe_print(f"  Admin:    http://127.0.0.1:{port}/admin")
        safe_print(f"  Account:  admin / admin123")
        safe_print("-" * 50)
        safe_print("  Press Ctrl+C to stop")
        safe_print("=" * 50)

        _crash_log("init: importing uvicorn...")
        import uvicorn
        _crash_log(f"init: starting server on {host}:{port}")

        # 非 daemon 线程跑 uvicorn，确保不会被主线程退出杀掉
        import threading as _th
        server_ready = _th.Event()

        def run_server():
            cfg = uvicorn.Config(app, host=host, port=port, log_level="warning")
            srv = uvicorn.Server(cfg)
            server_ready.set()
            srv.run()

        server_thread = _th.Thread(target=run_server, daemon=False)
        server_thread.start()

        # 等服务器就绪后打开浏览器
        server_ready.wait(timeout=3)
        import time
        time.sleep(0.8)  # 给 uvicorn 一点时间真正绑定端口

        url = f"http://127.0.0.1:{port}"
        public_url = None

        # --share: 启动 ngrok 公网隧道
        if use_share:
            _crash_log("init: --share enabled, preparing ngrok tunnel...")
            safe_print("")
            safe_print("-" * 50)
            safe_print("  公网隧道 (ngrok)")
            safe_print("-" * 50)
            safe_print("  默认情况下，只有和你在同一个 WiFi / 局域网的人")
            safe_print("  才能访问这个排行榜（比如同一个路由器下的设备）。")
            safe_print("")
            safe_print("  开启公网隧道后，会生成一个 https 公网链接，发给")
            safe_print("  任何人（无论在哪、用什么网络）都能打开。")
            safe_print("")
            safe_print("  不开公网的话：同一个 WiFi 下照常访问，不影响。")
            safe_print("  开了公网的好处：外地/异网的朋友也能连进来。")
            safe_print("")
            safe_print("  借助 ngrok 实现（https://ngrok.com，合法工具）。")
            safe_print("")

            do_connect = False
            try:
                import shutil as _shutil
                from pyngrok import ngrok, conf

                # 检查 ngrok 程序是否已安装（PATH 或常见位置）
                ngrok_path = _shutil.which("ngrok") or _shutil.which("ngrok.exe")
                if not ngrok_path:
                    # 也查 pyngrok 的默认安装目录
                    ngrok_default = os.path.join(os.path.expanduser("~"), ".pyngrok")
                    # pyngrok 内部路径: ~/.pyngrok/ngrok 或 ngrok.exe
                    for _ext in ("", ".exe"):
                        _p = os.path.join(ngrok_default, "ngrok" + _ext)
                        if os.path.exists(_p):
                            ngrok_path = _p
                            break

                if ngrok_path:
                    safe_print("  检测到 ngrok 已安装，直接启动隧道...")
                    do_connect = True
                else:
                    safe_print("  未检测到 ngrok，首次使用将自动下载 (~10MB)。")
                    safe_print("  下载后安装到用户目录，仅本程序使用。")
                    safe_print("")
                    safe_print("  如不放心，可手动安装后重试：")
                    safe_print("    https://ngrok.com/download")
                    safe_print("")
                    safe_print("-" * 50)
                    print("    按 Enter 同意下载，输入 n 取消 > ", end="", flush=True)
                    try:
                        choice = input().strip().lower()
                    except (EOFError, KeyboardInterrupt):
                        choice = "n"
                    if choice in ("n", "no"):
                        safe_print("    已取消。就可用局域网模式（同 WiFi 下照常访问）。")
                        _crash_log("init: user declined ngrok download")
                    else:
                        do_connect = True
                        safe_print("    正在下载 ngrok... 请稍候。")
                        _crash_log("init: user accepted, downloading ngrok...")
            except ImportError:
                _crash_log("init: pyngrok not installed")
                safe_print("  pyngrok 库未找到，无法使用 --share。")
                safe_print("  可就近访问（局域网内仍可连接）。")
                safe_print("-" * 50)
                safe_print("")
            except Exception as _ne:
                _crash_log(f"init: ngrok check failed: {_ne}")
                safe_print(f"  检查失败: {_ne}")
                safe_print("  可就近访问（局域网内仍可连接）。")
                safe_print("-" * 50)
                safe_print("")

            if do_connect:
                try:
                    ngrok_token = os.environ.get("NGROK_AUTHTOKEN", "")
                    if ngrok_token:
                        conf.get_default().auth_token = ngrok_token
                    public_url = ngrok.connect(port, "http").public_url
                    _crash_log(f"init: ngrok tunnel OK -> {public_url}")
                    safe_print("")
                    safe_print(f"  公网地址: {public_url}")
                    safe_print("  把这个链接发给任何人即可访问！")
                    safe_print("-" * 50)
                    safe_print("")
                except Exception as _ne:
                    _crash_log(f"init: ngrok failed: {_ne}")
                    safe_print(f"  [错误] 隧道启动失败: {_ne}")
                    safe_print("  可就近访问（局域网内仍可连接）。")
                    safe_print("-" * 50)
                    safe_print("")

        if open_browser:
            _crash_log("init: server ready, opening browser...")
            _crash_log(f"init: opening browser -> {url}")
            safe_print(f"  Opening browser: {url}")

            if browser_path:
                # 用户指定了浏览器
                import subprocess
                _crash_log(f"init: launching custom browser: {browser_path}")
                try:
                    subprocess.Popen([browser_path, url])
                    _crash_log("init: custom browser OK")
                except Exception as _be:
                    _crash_log(f"init: custom browser failed: {_be}")
                    safe_print(f"  Browser failed, trying default...")
                    os.startfile(url)
            else:
                # 系统默认浏览器
                try:
                    os.startfile(url)
                    _crash_log("init: os.startfile OK")
                except Exception as _be:
                    _crash_log(f"init: os.startfile failed: {_be}")
                    import webbrowser
                    webbrowser.open(url)
        else:
            _crash_log("init: --no-browser, skipping browser")
            safe_print(f"  Open manually: {url}")

        # 主线程等待
        try:
            while server_thread.is_alive():
                server_thread.join(1)
        except KeyboardInterrupt:
            _crash_log("init: KeyboardInterrupt, shutting down")
            safe_print("\n  Stopped.")

    except KeyboardInterrupt:
        safe_print("\n  Stopped.")
    except Exception as e:
        _crash_log(f"FATAL: {e}")
        _crash_log(_tb.format_exc())
        safe_print("")
        safe_print("=" * 50)
        safe_print("  ERROR: Failed to start!")
        safe_print("=" * 50)
        safe_print(f"  {e}")
        try:
            tb = _tb.format_exc()
            safe_print(tb)
        except Exception:
            safe_print(str(e))
        safe_print("-" * 50)
        safe_print("  Tips:")
        safe_print("  1. Port 8000 may be in use")
        safe_print("  2. Antivirus may be blocking")
        safe_print("  3. See crash.log for details")
        safe_print("=" * 50)
        exit_code = 1
    finally:
        if exit_code != 0:
            try:
                input("\n  Press Enter to exit...")
            except (EOFError, KeyboardInterrupt):
                pass

    sys.exit(exit_code)





